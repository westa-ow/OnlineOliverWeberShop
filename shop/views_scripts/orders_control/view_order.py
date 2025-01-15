import ast
import random
from datetime import datetime
from random import randint

import openpyxl
import concurrent.futures
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout, get_user_model, update_session_auth_hash
import os
import json
import firebase_admin
from django.views.decorators.csrf import csrf_exempt
from firebase_admin import credentials, firestore
from django.conf import settings
from django.contrib import messages
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.mail import send_mail
from openpyxl.reader.excel import load_workbook

from shop.forms import UserRegisterForm, User
from shop.views import addresses_ref, cart_ref, get_user_category, serialize_firestore_document, users_ref, is_admin, \
    orders_ref, itemsRef, db, process_items, get_user_prices, single_order_ref
from shop.views import get_user_info


@login_required
@user_passes_test(is_admin)
def view_order(request, order_id):
    chosenOrderRef = orders_ref.where("`order-id`", '==', int(order_id)).limit(1).stream()
    specificOrderData = {}
    specificOrderRef = ""
    found = False
    for chosenReference in chosenOrderRef:
        specificOrderData = chosenReference.to_dict()
        specificOrderRef = chosenReference
        found = True
    if not found:
        # If no results, perform a fallback query based on another condition, such as trying just 'order_id'
        # Assuming the fallback condition is meant to use a slightly different or less strict query condition
        fallbackOrderRef = orders_ref.where("order_id", '==', int(order_id)).limit(1).stream()
        for fallbackReference in fallbackOrderRef:
            specificOrderData = fallbackReference.to_dict()
            specificOrderRef = fallbackReference
            break
            # Assuming you have a way to reference your 'Item' collection
    itemList = specificOrderData.get('list', [])

    order_items = []
    with concurrent.futures.ThreadPoolExecutor() as executor:
        future = executor.submit(process_items, itemList)
        try:
            order_items = future.result(timeout=30)  # Adding a generous timeout to see if it helps
        except Exception as e:
            print(f"Unhandled exception: {e}")

    user_email = specificOrderData['email']

    category, currency = get_user_prices(request,user_email)
    info = get_user_info(user_email)

    serialized_data = (serialize_firestore_document(specificOrderRef))
    del serialized_data['list']

    for item in order_items:
        item['image_url'] = item['image_url'] if 'image_url' in item else item['image-url']
    context = {
        'feature_name': "view_order",
        'currency': "€" if currency == "Euro" else "$",
        'cart': [],
        'addresses': [],
        'user_info': info,
        'order_id': serialized_data['order-id'] or serialized_data['order_id'],
        'user_orders': order_items,  # Add order items to context
        'Order': serialized_data,
    }
    return render(request, 'admin_tools.html', context)

@login_required
@user_passes_test(is_admin)
def change_in_stock(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            orderId = data.get('orderId')
            productId = data.get('productId')
            stock_value = data.get('new_stock_value')

            query = single_order_ref.where("`order-id`", '==', int(orderId)).where('name', '==', productId).limit(1).stream()
            found = False
            for item in query:
                document_ref = item.reference
                document_ref.update({'in_stock': stock_value})
                found = True
            if not found:
                fallbackOrderRef = single_order_ref.where("order_id", '==', int(orderId)).where('name', '==', productId).limit(1).stream()
                for item in fallbackOrderRef:
                    document_ref = item.reference
                    document_ref.update({'in_stock': stock_value})
                    break
            return JsonResponse({"success": True, "message": "Order 'in stock' updated successfully."})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@login_required
def change_tracker_link(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            orderId = data.get('orderId')
            tracker_link = data.get('tracker')

            query = orders_ref.where("`order-id`", '==', int(orderId)).limit(1).stream()
            found = False
            for item in query:
                document_ref = item.reference
                document_ref.update({'tracker': tracker_link})
                found = True
            if not found:
                fallbackOrderRef = orders_ref.where("order_id", '==', int(orderId)).limit(1).stream()
                for item in fallbackOrderRef:
                    document_ref = item.reference
                    document_ref.update({'tracker': tracker_link})
                    break
            return JsonResponse({"success": True, "message": "Order tracker updated successfully."})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@login_required
@user_passes_test(is_admin)
def upload_in_stock(request, order_id):
    if request.method == 'POST':
        xlsx_file = request.FILES['xlsx_file']

        if xlsx_file.name.endswith('.xlsx'):
            wb = load_workbook(filename=xlsx_file, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
                product_name = str(row[0])
                in_stock_indicator = row[1]
                in_stock = bool(in_stock_indicator)  # Converts 1 or 0 to True or False

                # Query database to find documents to update
                query = single_order_ref.where("`order-id`", '==', int(order_id)).where('name', '==',
                                                                                        str(product_name)).limit(1).stream()
                found = False
                for item in query:
                    document_ref = item.reference
                    document_ref.update({'in_stock': in_stock})
                    found = True
                if not found:
                    fallbackOrderRef = single_order_ref.where("order_id", '==', int(order_id)).where('name', '==', product_name).limit(1).stream()
                    for item in fallbackOrderRef:
                        document_ref = item.reference
                        document_ref.update({'in_stock': in_stock})
                        break
            return JsonResponse({"success": True, "message": "Order 'in stock' updated successfully."}, status=200)
        else:
            return JsonResponse({'status': 'error', 'message': "Invalid file format"}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)
