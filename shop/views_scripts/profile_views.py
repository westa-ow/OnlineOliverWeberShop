import concurrent

from django.contrib.auth.decorators import login_required
from django.core.files.storage import FileSystemStorage
from google.cloud.firestore_v1 import DocumentReference
from openpyxl import load_workbook

import xml.etree.ElementTree as ET

from shop.views import db, orders_ref, serialize_firestore_document, users_ref, addresses_ref, update_email_in_db, \
    get_user_category, get_user_info, get_vocabulary_product_card, get_user_prices, get_user_session_type, itemsRef, \
    get_user_sale
import ast
import random
from datetime import datetime
from random import randint

import concurrent.futures
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout, get_user_model, update_session_auth_hash
import os
import json
import firebase_admin
from django.views.decorators.csrf import csrf_exempt
from firebase_admin import credentials, firestore
from django.conf import settings
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.mail import send_mail

from django.utils.translation import gettext as _

from shop.views_scripts.catalog_views import update_cart, get_full_product


@login_required
def profile(request, feature_name):
    """
    This function is used to render the main profile page for the user. It takes feature_name parameter and renders corresponding template.
    :param request:
    :param feature_name:
    :return:
    """
    email = request.user.email
    orders = get_orders_for_user(email)
    order_details = get_order_details(orders)
    email = request.user.email
    category, currency = get_user_prices(request,email)
    info = get_user_info(email) or {}
    show_quantities = info['show_quantities'] if 'show_quantities' in info else False
    if currency == "Euro":
        currency = "€"
    elif currency == "Dollar":
        currency = "$"

    if feature_name == 'order_import' or feature_name == "archive_search":
        if info.get('customer_type', 'Customer') != 'B2B':
            return redirect('profile', feature_name='dashboard')
    context = build_context(feature_name, email, orders, order_details)
    context['currency'] = currency
    context['userId'] = info['userId']
    context['username'] = info['first_name'] + " " + info['last_name']
    context['show_quantities'] = show_quantities
    context['STRIPE_PUBLISHABLE_KEY'] = settings.STRIPE_PUBLISHABLE_KEY

    if feature_name == "catalogs":
        pdf_files = [
            {
                'name': 'Main Catalog 2024',
                'preview_img_url': 'https://firebasestorage.googleapis.com/v0/b/flutterapp-fd5c3.appspot.com/o/website%2Fcatalogs_previews%2Fmain2024_preview.png?alt=media&token=ec4e8220-76a2-4e28-95b6-a4b014bc3ef5',
                'url': 'https://firebasestorage.googleapis.com/v0/b/flutterapp-fd5c3.appspot.com/o/website%2Fcatalogs%2FMain%20Catalog%20with%20Price%20NO%20barcodes%20LOW.pdf?alt=media&token=d2d7146e-30f8-4865-8875-f72ef36dc1cc'
            },
            {
                'name': 'Fortune, Fame, Love',
                'preview_img_url': 'https://firebasestorage.googleapis.com/v0/b/flutterapp-fd5c3.appspot.com/o/website%2Fcatalogs_previews%2Ffortunelove_preview.png?alt=media&token=2b47d026-e775-41d8-ba83-512d3015f5aa',
                'url': 'https://firebasestorage.googleapis.com/v0/b/flutterapp-fd5c3.appspot.com/o/website%2Fcatalogs%2FFortune_Fame_Love.pdf?alt=media&token=31084978-b337-40ea-aa75-28e23cb5385f'
            },
            {
                'name': 'Summer Collection 2024',
                'preview_img_url': 'https://firebasestorage.googleapis.com/v0/b/flutterapp-fd5c3.appspot.com/o/website%2Fcatalogs_previews%2Fsummer2024_preview.png?alt=media&token=18982b88-27dd-41fe-a617-58df02e9b092',
                'url': 'https://firebasestorage.googleapis.com/v0/b/flutterapp-fd5c3.appspot.com/o/website%2Fcatalogs%2FCatalog%20Summer%202024_with%20Price%20HIGH.pdf?alt=media&token=1086d001-13c2-4f02-803b-298d424013a1'
            },
            {
                'name': 'Showcase & Display',
                'preview_img_url': 'https://firebasestorage.googleapis.com/v0/b/flutterapp-fd5c3.appspot.com/o/website%2Fcatalogs_previews%2Fshowcase2024_preview.png?alt=media&token=9456eb76-f230-48d5-92d1-9048ad4ca911',
                'url': 'https://firebasestorage.googleapis.com/v0/b/flutterapp-fd5c3.appspot.com/o/website%2Fcatalogs%2FShowcase%20%26%20Display%20Spirits%202024.pdf?alt=media&token=29e6fea9-a93b-4fe6-844d-26a388811eb3'
            },

            {
                'name': 'Main Catalog 2025',
                'preview_img_url': 'https://firebasestorage.googleapis.com/v0/b/flutterapp-fd5c3.appspot.com/o/website%2Fcatalogs_previews%2Fmainpreview.png?alt=media&token=e26efe00-5db1-47c8-bdcf-721651aa5125',
                'url': 'https://firebasestorage.googleapis.com/v0/b/flutterapp-fd5c3.appspot.com/o/website%2Fcatalogs%2FHauptkatalog%202025_without%20price_compressed.pdf?alt=media&token=7a6a350d-34fd-466f-847f-391f021c1480'
            },
            {
                'name': 'Winter Catalog 2025',
                'preview_img_url': 'https://firebasestorage.googleapis.com/v0/b/flutterapp-fd5c3.appspot.com/o/website%2Fcatalogs_previews%2Fwinter2025_preview.png?alt=media&token=6f71dc5e-0128-4b2a-bbd5-30f072d08d85',
                'url': 'https://firebasestorage.googleapis.com/v0/b/flutterapp-fd5c3.appspot.com/o/website%2Fcatalogs%2FCatalog%20NEW%20COLLECTION2025_WITHOUT%20PRICE_compressed.pdf?alt=media&token=2b2ea167-9dfa-40e1-94f4-bcecac36970c'
            },
            {
                'name': 'Spring Catalog 2025',
                'preview_img_url': 'https://firebasestorage.googleapis.com/v0/b/flutterapp-fd5c3.appspot.com/o/website%2Fcatalogs_previews%2Fspringpreview.png?alt=media&token=154458a3-92e6-470c-87d9-1ff1dbdf3ddd',
                'url': 'https://firebasestorage.googleapis.com/v0/b/flutterapp-fd5c3.appspot.com/o/website%2Fcatalogs%2FCatalog%20SPRING%202025%20WITH%20price%20PRINT.pdf?alt=media&token=756d9c51-d9c1-46c8-bd79-b3404eb99e05'
            },

        ]
        context['pdf_files'] = pdf_files

    return render(request, 'profile.html', context=context)


def get_orders_for_user(email):
    """
    Get all orders for a specific user by email.
    :param email:
    :return:
    """
    orders = []
    docs_orders = orders_ref.where('email', '==', email).stream()

    for doc in docs_orders:
        order_info = doc.to_dict()
        order_id = order_info.get('order_id') or order_info.get('order-id')
        formatted_date = format_date(order_info.get('date'))
        orders.append({
            'Status': order_info.get('Status'),
            'date': formatted_date,
            'email': email,
            'list': order_info.get('list'),
            'order_id': order_id,
            'sum': order_info.get('price'),
            'shippingPrice': order_info.get('shippingPrice', 0),
            'currency': "€" if order_info.get('currency') == "Euro" else "$",
            'payment_type': order_info.get('payment_type', "BANK TRANSFER"),
            'tracker': order_info.get('tracker', 'None'),
            'paid_sum': order_info.get('paid_sum', 0),
            'shippingAddressId': order_info.get('shippingAddressId', ""),
            'billingAddressId': order_info.get('billingAddressId', ""),
        })
    orders.sort(key=lambda x: x['date'], reverse=True)
    return orders

def make_json_serializable(obj):
    if isinstance(obj, DocumentReference):
        # choose whichever you need on the client side:
        return obj.path       # or obj.id
    elif isinstance(obj, dict):
        return {k: make_json_serializable(v) for k, v in obj.items()}
    elif isinstance(obj, (list, tuple)):
        return [make_json_serializable(v) for v in obj]
    else:
        return obj

def format_date(date_obj):
    """
    Format date object to string in specific format.
    :param date_obj:
    :return:
    """
    return date_obj.strftime("%Y-%m-%d") if date_obj else None


def get_order_details(orders):
    """
    Fetch order details for each order in the list of orders.
    :param orders:
    :return:
    """
    order_details = {order['order_id']: [] for order in orders}

    with concurrent.futures.ThreadPoolExecutor() as executor:
        future_to_order = schedule_order_detail_fetching(executor, orders, order_details)
        for future in concurrent.futures.as_completed(future_to_order):
            order_id = future_to_order[future]
            order_doc_data = future.result()
            if order_doc_data:
                order_details[order_id].append(order_doc_data)

    return order_details


def schedule_order_detail_fetching(executor, orders, order_details):
    """
    Schedule fetching of order details for each order in the list of orders.
    :param executor:
    :param orders:
    :param order_details:
    :return:
    """
    future_to_order = {}
    for order in orders:
        for order_doc_ref in order['list']:
            order_doc_path = order_doc_ref if type(order_doc_ref) is str else order_doc_ref.path
            future = executor.submit(fetch_order_detail, order_doc_path)
            future_to_order[future] = order['order_id']
    return future_to_order


def fetch_order_detail(order_doc_path):
    """
    Fetch order detail for a specific order path.
    :param order_doc_path:
    :return:
    """
    path_parts = order_doc_path.split('/')
    if len(path_parts) == 2:
        collection_name, document_id = path_parts
        order_doc_ref = db.collection(collection_name).document(document_id)
        order_doc = order_doc_ref.get()
        if order_doc.exists:
            return order_doc.to_dict()
    return None


def build_context(feature_name, email, orders, order_details):
    """
    Build context for the profile page for the specific profile page.
    :param feature_name:
    :param email:
    :param orders:
    :param order_details:
    :return:
    """
    currencies_dict ={}
    for order in orders:
        currencies_dict[order['order_id']] = order['currency']
    config = {
        "orders": orders,
        "currencies": currencies_dict,
        "products": order_details
    }
    context = {
        'currencies': currencies_dict,
        'orders': orders,
        'products': order_details,
        'feature_name': feature_name,
        'vocabulary': get_vocabulary_product_card()
    }

    if feature_name == "account":
        context['user_info'], context['user_info_dict'] = get_user_info_dicts(email)

    if feature_name == "addresses":
        addresses, addresses_dict = get_user_addresses(email)
        context['my_addresses'] = addresses
        context['addresses_dict'] = addresses_dict
        config['my_addresses'] = addresses
        config['addresses_dict'] = addresses_dict
    config_data = make_json_serializable(config)
    context['config_data'] = config_data
    return context


def get_user_info_dicts(email):
    """
    Fetch user information from Firestore and convert it to dictionaries.
    :param email:
    :return:
    """
    existing_users = users_ref.where('email', '==', email).limit(1).get()
    if existing_users:
        for user in existing_users:
            user_ref = users_ref.document(user.id)
            user_data = serialize_firestore_document(user_ref.get())
            information = json.dumps(user_data)
            information2 = json.loads(information)
            return information2, information
    return None, None


def get_user_addresses(email):
    """
    Fetch user addresses from Firestore and convert it to dictionaries.
    :param email:
    :return:
    """
    addresses = []
    existing_addresses = addresses_ref.where('email', '==', email).get()
    if existing_addresses:
        for address in existing_addresses:
            address_doc = addresses_ref.document(address.id).get().to_dict()
            if address_doc.get('creation_date', '') != '':
                address_doc['creation_date'] = address_doc['creation_date'].strftime('%Y-%m-%d')
            if address_doc.get('is_deleted', False) != True:
                addresses.append(address_doc)


    addresses_dict = json.dumps(addresses)
    return addresses, addresses_dict



@login_required
def update_user_account(request):
    """
    Update user account information. This function is called when the user updates their account information.
    :param request:
    :return:
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            old_data = data.get('old')
            new_data = data.get('new')

            old_email = old_data['email']
            new_email = new_data['email']

            # Check for existing user by email
            existing_users = users_ref.where('email', '==', old_email).limit(1).get()
            if new_data['firstname'] != old_data['first_name'] or new_data['lastname'] != old_data['last_name']:
                User = get_user_model()
                try:
                    user_instance = User.objects.get(email=old_email)
                    user_instance.first_name = new_data['firstname']
                    user_instance.last_name = new_data['lastname']
                    user_instance.save()
                except User.DoesNotExist:
                    return JsonResponse({'status': 'error', 'message': 'User not found.'}, status=404)

            if new_data['email'] != old_email:
                existing_user_with_new_email = users_ref.where('email', '==', new_email).limit(1).get()
                if len(existing_user_with_new_email) > 0:
                    return JsonResponse({'status': 'error', 'message': 'User with this email exists.'}, status=400)
                User = get_user_model()
                try:
                    user_instance = User.objects.get(email=old_email)
                    user_instance.email = new_email  # Update the email
                    user_instance.username = new_email  # Update the email
                    user_instance.save()
                except User.DoesNotExist:
                    return JsonResponse({'status': 'error', 'message': 'User not found.'}, status=404)

                try:
                    update_email_in_db(old_email, new_email)
                    # Optionally use update_email_in_db_result for further logic
                except Exception as e:
                    # Log the exception e
                    return JsonResponse(
                        {'status': 'error', 'message': 'An error occurred while updating documents in Firestore.'},
                        status=500)
            else:
                User = get_user_model()
                try:
                    # Retrieve the user instance
                    user_instance = User.objects.get(email=old_email)
                except User.DoesNotExist:
                    return JsonResponse({'status': 'error', 'message': 'User not found.'}, status=404)

            social_title = old_data['social_title'] if 'id_gender' not in new_data else "Mr" if new_data[
                                                                                                    'id_gender'] == "1" else "Mrs"
            receive_newsletter = old_data['receive_newsletter'] if 'newsletter' not in new_data else True if new_data[
                                                                                                                 'newsletter'] == "1" else False
            receive_offers = old_data['receive_offers'] if 'optin' not in new_data else True if new_data[
                                                                                                    'optin'] == "1" else False
            password = new_data['password']

            # Check if the provided password matches the user's password
            if not user_instance.check_password(password):
                return JsonResponse({'status': 'error', 'message': 'Incorrect password.'}, status=400)

            new_password = new_data.get('new_password', '')
            if new_password:  # This checks if the new_password string is not empty
                user_instance.set_password(new_password)
                user_instance.save()  # Don't forget to save the user object after setting the new password
                update_session_auth_hash(request, user_instance)

            user_instance.save()

            for user in existing_users:
                user_ref = users_ref.document(user.id)
                user_ref.update({
                    'social_title': social_title,
                    'first_name': new_data['firstname'],
                    'last_name': new_data['lastname'],
                    'email': new_data['email'],
                    'birthday': new_data['birthday'],
                    'receive_newsletter': receive_newsletter,
                    'receive_offers': receive_offers,
                })
                return JsonResponse({'status': 'success', 'message': 'User updated successfully.'})

            return JsonResponse({'status': 'success', 'message': 'User account updated successfully.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=400)


def upload_file(request):
    """
    Uploads a file with products that will be added to user's cart.
    :param request:
    :return:
    """
    if request.method == 'POST' and 'file' in request.FILES:
        uploaded_file = request.FILES['file']

        errors = []

        # Checking for correct file extension
        if not (uploaded_file.name.endswith('.xlsx') or uploaded_file.name.endswith('.xls')):
            errors.append(_('Invalid file format. Only .xlsx or .xls are allowed'))
            return JsonResponse({'status': 'error', 'message': 'Invalid file format. Only .xlsx or .xls are allowed', 'errors': errors}, status=400)

        try:
            # Open the downloaded file with openpyxl
            workbook = load_workbook(uploaded_file)
            sheet = workbook.active

            # Go through the lines of the file and call the required function
            for row in sheet.iter_rows(min_row=2, values_only=True):
                product_name, new_quantity = row
                if not product_name or new_quantity is None:
                    errors.append(_('Invalid data in row') + f': {row}')
                    continue  # Skip rows with missing data

                product_name = str(product_name)
                new_quantity = int(new_quantity)
                # Get data from the user session and call the necessary functions
                email = get_user_session_type(request)
                category, currency = get_user_prices(request, email)
                info = get_user_info(email) or {}
                sale = get_user_sale(info)

                # Check for product name and quantity
                if not product_name or new_quantity is None:
                    errors.append(_('Invalid data in row') + f': {row}')
                    continue  # Skip rows without product or quantity
                if new_quantity <= 0:
                    errors.append(_("Quantity can only be greater than 0)") + f": {row}")
                    continue
                # Looking for product information
                document = get_full_product(product_name)
                if not document:
                    errors.append(_('Product not found') + f': {product_name}')
                    continue  # If no product is found, continue processing the next line

                if document['quantity'] < new_quantity:
                    errors.append(_('Requested quantity') + f'({new_quantity})' + _('is less than number of products in storage for product')+ f': {product_name}')
                    continue
                # Determine the price of the product depending on the category
                if category == "VK3":
                    document['price'] = document.get('priceVK3', 0)
                elif category == "GH":
                    document['price'] = document.get('priceGH', 0)
                elif category == "Default_High":
                    document['price'] = document.get('priceVK4', 0) * 1.3
                elif category == "Default_USD":
                    document['price'] = round(document.get('priceUSD', 0) * (1-sale), 1) or 0
                elif category == "GH_USD":
                    document['price'] = document.get('priceUSD_GH', 0)
                else:
                    document['price'] = round(document.get('priceVK4', 0) * (1-sale), 1) or 0

                # Update the cart
                subtotal, cart_size = update_cart(email, product_name, new_quantity, document)

                if subtotal is None:
                    errors.append(_('Error updating cart for product') + f': {product_name}')
                    return JsonResponse({'status': 'error', 'message': 'An error occurred while processing your request', 'errors': errors}, status=500)

            return JsonResponse({'status': 'success', 'message': 'All products processed successfully', 'errors': errors}, status=200)

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Error processing file: {str(e)}', 'errors': errors}, status=500)
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)


def generate_product_feed(request):
    """
    This function generates xml product feed.
    :param request:
    :return:
    """
    products = itemsRef.stream()

    root = ET.Element("products")

    for product in products:
        data = product.to_dict()
        if not data.get("Visible", False) or int(data.get('quantity', 0)) <= 0:
            continue

        product_element = ET.SubElement(root, "product")
        ET.SubElement(product_element, "name").text = f"{data.get('category', '')} {data.get('product_name', '')}"
        ET.SubElement(product_element, "article_number").text = data.get("name", "")
        ET.SubElement(product_element, "ean_13").text = data.get("ean_13", "")

        ET.SubElement(product_element, "availability").text = "in stock" if data.get("quantity", 0) > 0 else "Out of stock"
        ET.SubElement(product_element, "image").text = data.get("image_url", "https://firebasestorage.googleapis.com/v0/b/flutterapp-fd5c3.appspot.com/o/wall%2Fno_image.jpg?alt=media&token=22a7b907-01f6-45b6-8fb1-f1f884ab21d4")
        ET.SubElement(product_element, "quantity").text = str(data.get("quantity", 0))
        if data.get("product_width"):
            ET.SubElement(product_element, "width").text = str(data.get("product_width", "")) + " cm"
        if data.get("product_height"):
            ET.SubElement(product_element, "height").text = str(data.get("product_height", "")) + " cm"
        if data.get("chain_length"):
            ET.SubElement(product_element, "length").text = str(data.get("chain_length", "")) + " cm"


    xml_data = ET.tostring(root, encoding="utf-8")

    return HttpResponse(xml_data, content_type="application/xml")