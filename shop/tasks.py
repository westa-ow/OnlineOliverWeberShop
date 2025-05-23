import os

import openpyxl
import csv
import io
import firebase_admin
from firebase_admin import credentials, firestore
import datetime
from background_task import background

from OnlineShop import settings
from OnlineShop.settings import FIREBASE_CREDENTIALS
from django.core.mail import send_mail, EmailMessage

@background(schedule=30)
def process_file_task(xlsx_file_path):

    logs = []  # log storage list

    def log(message):
        logs.append(f"{datetime.datetime.now().isoformat()} - {message}")
        print(message)

    try:
        firebase_admin.get_app()
        log("Firebase app already initialized.")
    except ValueError:
        firebase_admin.initialize_app(FIREBASE_CREDENTIALS)
        log("Firebase app initialized.")

    collection_ref = firestore.client().collection("item")

    # Download the Excel file
    try:
        workbook = openpyxl.load_workbook(xlsx_file_path)
        log(f"Workbook '{xlsx_file_path}' loaded successfully.")
    except Exception as e:
        log(f"Error loading workbook: {e}")
        send_email_with_logs(logs)
        return
    sheet = workbook.active

    # Determine the numbers of the required columns
    columnQuantity = None
    columnNumber = None
    for col_index, col in enumerate(sheet.iter_cols()):
        header_value = col[0].value
        if header_value == 'Lagerstand Gesamt':
            columnQuantity = col_index
            log(f"Found 'Lagerstand Gesamt' at column index {col_index}.")
        elif header_value == 'Artikelnummer':
            columnNumber = col_index
            log(f"Found 'Artikelnummer' at column index {col_index}.")

    if columnQuantity is None or columnNumber is None:
        log("Required column headers not found.")
        send_email_with_logs(logs)
        return

    # Create in-memory CSV
    output = io.StringIO()
    csv_writer = csv.writer(output, delimiter=';')
    csv_writer.writerow(['name', 'quantity'])
    log("CSV header written.")

    # Processing file lines
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[columnNumber]
        quantity = row[columnQuantity] if row[columnQuantity] is not None else 0
        try:
            quantity = max(int(round(quantity, 0)), 0)
        except Exception as e:
            log(f"Error processing quantity for {name}: {e}")
            quantity = 0
        csv_writer.writerow([name, quantity])
    log("CSV data prepared.")

    output.seek(0)

    # Batch update Firestore
    batch = firestore.client().batch()
    csv_reader = csv.reader(output, delimiter=';')
    next(csv_reader)  # Skip the header
    for row in csv_reader:
        query_value = row[0]
        try:
            new_value = int(row[1])
        except ValueError:
            new_value = int(round(float(row[1])))
        query = collection_ref.where("name", "==", query_value)
        query_results = query.get()
        for doc_snapshot in query_results:
            field1_value = doc_snapshot.get("quantity")
            if field1_value != new_value:
                log(f"Product: {doc_snapshot.get('name')}    Old quantity: {field1_value}    New quantity: {new_value}")
                batch.update(doc_snapshot.reference, {"quantity": new_value})
    try:
        batch.commit()
        log("Firestore update was successful.")
    except Exception as e:
        log(f"Error occurred during Firestore update: {e}")
    print("Database was successfully updated!")
    log("Database was successfully updated!")
    send_email_with_logs(logs)


def send_email_with_logs(logs):
    """
    Writes logs to a text file and sends an email with an attachment.
    """
    log_text = "\n".join(logs)
    # Define the name of the log file
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"processing_log_{timestamp}.txt"
    file_path = os.path.join(settings.BASE_DIR, filename)

    # Writing logs to a file
    try:
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(log_text)
    except Exception as e:
        print(f"Error writing log file: {e}")
        return

    # Формируем email
    subject = "Db update report"
    body = "Attached file is the log of the database update task execution."
    from_email = settings.EMAIL_HOST_USER
    recipient_list = ["eramcheg@gmail.com", "westadatabase@gmail.com"]

    email = EmailMessage(subject, body, from_email, recipient_list)
    email.attach(filename, log_text, "text/plain")

    try:
        email.send()
        print("Email sent successfully with logs.")
    except Exception as e:
        print(f"Error sending email: {e}")